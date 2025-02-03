import streamlit as st
import pandas as pd
from datetime import datetime
import re
import io
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import zipfile
import os
from pathlib import Path
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ページ設定
st.set_page_config(
    page_title="PINO精算アプリケーション",
    layout="wide"
)

# 定数
RATE_PER_KM = 15
DAILY_ALLOWANCE = 200

# スタイル
st.markdown("""
    <style>
    .main {
        padding: 20px;
    }
    .stTextArea textarea {
        font-size: 16px;
    }
    .expense-table {
        font-size: 14px;
    }
    .stButton>button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        padding: 8px 16px;
        border: none;
        border-radius: 4px;
    }
    .expense-table th {
        text-align: center !important;
    }
    .expense-table td {
        text-align: right !important;
    }
    .expense-table td:nth-child(2) {
        text-align: left !important;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 16px;
        background-color: #f0f2f6;
    }
    .stTabs [aria-selected="true"] {
        background-color: #4CAF50 !important;
        color: white !important;
    }
    </style>
""", unsafe_allow_html=True)

def parse_expense_data(text):
    """テキストデータを解析してDataFrameに変換"""
    try:
        # テキストを行に分割
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # データを格納するリスト
        data = []
        
        # 各行を解析
        current_name = None
        for line in lines:
            # 名前の行を検出
            if '様' in line:
                current_name = line.replace('様', '').strip()
                continue
            
            # データ行を解析
            match = re.match(r'(\d{1,2}/\d{1,2})\s+(.+)', line)
            if match and current_name:
                date, route = match.groups()
                
                # 距離を計算（仮の実装）
                distance = len(route.split('→')) * 5.0  # 仮の距離計算
                
                # 交通費と手当を計算
                transportation_fee = distance * RATE_PER_KM
                allowance = DAILY_ALLOWANCE
                total = transportation_fee + allowance
                
                data.append({
                    'name': current_name,
                    'date': date,
                    'route': route,
                    'total_distance': distance,
                    'transportation_fee': transportation_fee,
                    'allowance': allowance,
                    'total': total
                })
        
        # DataFrameを作成
        if data:
            df = pd.DataFrame(data)
            return df
        else:
            st.error("有効なデータが見つかりませんでした。")
            return None
            
    except Exception as e:
        st.error(f"データの解析中にエラーが発生しました: {str(e)}")
        return None

def format_number(val):
    if pd.isna(val):
        return ''
    if isinstance(val, str) and '+' in val:  # 計算式の場合
        return val
    if isinstance(val, float):
        if val.is_integer():
            return f"{int(val):,}"
        return f"{val:.1f}"
    if isinstance(val, (int, str)):
        try:
            return f"{int(val):,}"
        except (ValueError, TypeError):
            return val
    return val

def create_expense_table_image(df, name, start_date):
    # フォントとサイズの設定
    title_font_size = 48  # さらに大きく
    header_font_size = 32
    content_font_size = 24
    padding = 80
    row_height = 80
    col_widths = [120, 600, 140, 160, 140, 140]  # 列幅をさらに広く
    
    # 画像サイズを大きくして高解像度に対応
    width = sum(col_widths) + padding * 2
    height = (len(df) + 3) * row_height + padding * 3
    
    # 画像の作成（サイズを3倍に）
    scale_factor = 3
    img = Image.new('RGB', (int(width * scale_factor), int(height * scale_factor)), 'white')
    draw = ImageDraw.Draw(img)
    
    # デフォルトフォントを使用
    title_font = ImageFont.load_default()
    header_font = ImageFont.load_default()
    content_font = ImageFont.load_default()
    
    def scale(x): return int(x * scale_factor)
    
    # タイトルの描画
    title = f"{name}様 1月 交通費清算書"
    draw.text((scale(padding), scale(padding)), title, fill='black', font=title_font)
    
    # ヘッダーの描画
    headers = ['日付', '経路', '距離\n(km)', '交通費\n(円)', '手当\n(円)', '合計\n(円)']
    x = scale(padding)
    y = scale(padding + row_height * 1.5)
    
    # ヘッダー背景
    for header, width in zip(headers, col_widths):
        draw.rectangle([x, y, x + scale(width), y + scale(row_height)], 
                      fill='#f5f5f5', outline='#666666', width=4)
        
        lines = header.split('\n')
        for i, line in enumerate(lines):
            # デフォルトフォントではgetlengthが使えないため、
            # 文字数に基づいて位置を計算
            text_width = len(line) * scale(header_font_size/2)
            text_x = x + (scale(width) - text_width) / 2
            text_y = y + scale(10) + (i * scale(row_height - 20) / len(lines))
            draw.text((text_x, text_y), line, fill='black', font=header_font)
        x += scale(width)
    
    # データの描画
    y += scale(row_height)
    for i, (_, row) in enumerate(df.iterrows()):
        x = scale(padding)
        for col_idx, (value, width) in enumerate(zip(row, col_widths)):
            draw.rectangle([x, y, x + scale(width), y + scale(row_height)], 
                         outline='#666666', width=4)
            
            text = str(value) if pd.notna(value) else ''
            
            if col_idx == 1:  # 経路列
                words = text.split('→')
                if len(words) >= 3:
                    mid_point = len(words) // 2
                    line1 = '→'.join(words[:mid_point]) + '→'
                    line2 = '→'.join(words[mid_point:])
                    
                    draw.text((x + scale(15), y + scale(10)), line1, 
                             fill='black', font=content_font)
                    draw.text((x + scale(15), y + scale(row_height/2)), line2, 
                             fill='black', font=content_font)
                else:
                    text_y = y + scale(row_height/2 - content_font_size/2)
                    draw.text((x + scale(15), text_y), text, 
                             fill='black', font=content_font)
            else:
                # 数値列は右寄せ（デフォルトフォントの場合）
                text_width = len(text) * scale(content_font_size/2)
                text_x = x + scale(width) - text_width - scale(15)
                text_y = y + scale(row_height/2 - content_font_size/2)
                draw.text((text_x, text_y), text, fill='black', font=content_font)
            x += scale(width)
        y += scale(row_height)
    
    # 注釈の描画
    note = "※2025年1月分給与にて清算しました。"
    draw.text((scale(padding), scale(height - row_height)), note, 
              fill='black', font=content_font)
    
    # 画像をバイト列に変換（300dpiで保存）
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG', dpi=(300, 300))
    img_byte_arr = img_byte_arr.getvalue()
    
    return img_byte_arr

def create_zip_file(images_dict):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for name, img_bytes in images_dict.items():
            zip_file.writestr(
                f"精算書_{name}_{datetime.now().strftime('%Y%m%d')}.png",
                img_bytes
            )
    return zip_buffer.getvalue()

def create_expense_report_pdf(df, name):
    # PDFバッファの作成
    buffer = io.BytesIO()
    
    # A4サイズの設定
    width, height = A4
    
    # PDFキャンバスの作成
    c = canvas.Canvas(buffer, pagesize=A4)
    
    # フォントの設定
    c.setFont('Helvetica', 12)
    
    # タイトルの描画
    c.setFont('Helvetica-Bold', 24)
    c.drawString(30*mm, 270*mm, f"{name}様 1月 交通費清算書")
    
    # ヘッダーの描画
    headers = ['日付', '経路', '距離(km)', '交通費(円)', '手当(円)', '合計(円)']
    col_widths = [25*mm, 70*mm, 20*mm, 25*mm, 25*mm, 25*mm]
    x_positions = [30*mm]
    for width in col_widths[:-1]:
        x_positions.append(x_positions[-1] + width)
    
    # ヘッダー背景
    c.setFillColorRGB(0.95, 0.95, 0.95)
    c.rect(30*mm, 250*mm, sum(col_widths), 10*mm, fill=1)
    
    # ヘッダーテキスト
    c.setFillColorRGB(0, 0, 0)
    c.setFont('Helvetica-Bold', 10)
    for header, x in zip(headers, x_positions):
        c.drawString(x + 2*mm, 252*mm, header)
    
    # データの描画
    c.setFont('Helvetica', 10)
    y_position = 240*mm
    for _, row in df.iterrows():
        # 罫線
        c.rect(30*mm, y_position - 8*mm, sum(col_widths), 10*mm)
        for x, width in zip(x_positions, col_widths):
            c.line(x, y_position - 8*mm, x, y_position + 2*mm)
        
        # データ
        values = [
            str(row['日付']),
            str(row['経路']),
            str(row['距離(km)']),
            str(row['交通費(円)']),
            str(row['手当(円)']),
            str(row['合計(円)'])
        ]
        
        for value, x, width in zip(values, x_positions, col_widths):
            if len(value) > 30:  # 経路が長い場合は2行に分割
                parts = value.split('→')
                mid = len(parts) // 2
                line1 = '→'.join(parts[:mid])
                line2 = '→'.join(parts[mid:])
                c.drawString(x + 2*mm, y_position - 2*mm, line1)
                c.drawString(x + 2*mm, y_position - 6*mm, line2)
            else:
                c.drawString(x + 2*mm, y_position - 4*mm, value)
        
        y_position -= 10*mm
    
    # 注釈の描画
    c.setFont('Helvetica', 10)
    c.drawString(30*mm, 30*mm, "※2025年1月分給与にて清算しました。")
    
    # PDFの保存
    c.save()
    buffer.seek(0)
    return buffer.getvalue()

def create_expense_excel(df, name):
    """Excelファイルを作成"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # データフレームをExcelに書き込み
        df.to_excel(writer, sheet_name='精算書', index=False, startrow=1)
        
        # ワークシートを取得
        ws = writer.sheets['精算書']
        
        # タイトルを追加
        ws.insert_rows(0)
        ws['A1'] = f"{name}様 1月 交通費清算書"
        ws.merge_cells('A1:F1')
        
        # スタイルの設定
        title_font = Font(size=14, bold=True)
        header_font = Font(size=11, bold=True)
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # タイトルのスタイル
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # ヘッダーのスタイル
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # データセルのスタイル
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 12  # 日付
        ws.column_dimensions['B'].width = 50  # 経路
        ws.column_dimensions['C'].width = 12  # 距離
        ws.column_dimensions['D'].width = 12  # 交通費
        ws.column_dimensions['E'].width = 12  # 手当
        ws.column_dimensions['F'].width = 12  # 合計
        
        # 注釈を追加
        last_row = len(df) + 3
        ws[f'A{last_row}'] = "※2025年1月分給与にて清算しました。"
        ws.merge_cells(f'A{last_row}:F{last_row}')
        
    return output.getvalue()

def main():
    st.title("PINO精算アプリケーション")
    
    input_text = st.text_area("精算データを貼り付けてください", height=200)
    
    if st.button("データを解析"):
        if input_text:
            df = parse_expense_data(input_text)
            if df is not None:
                st.session_state['expense_data'] = df
                st.success("データを解析しました！")
    
    if 'expense_data' in st.session_state:
        df = st.session_state['expense_data']
        unique_names = df['name'].unique().tolist()
        
        # タブの作成
        tabs = st.tabs(unique_names)
        
        for i, name in enumerate(unique_names):
            with tabs[i]:
                person_data = df[df['name'] == name].copy()
                if len(person_data) > 0:
                    styled_df = person_data[['date', 'route', 'total_distance', 'transportation_fee', 'allowance', 'total']]
                    styled_df.columns = ['日付', '経路', '距離(km)', '交通費(円)', '手当(円)', '合計(円)']
                    
                    # 数値のフォーマット
                    styled_df['距離(km)'] = styled_df['距離(km)'].map('{:.1f}'.format)
                    styled_df['交通費(円)'] = styled_df['交通費(円)'].map('{:,.0f}'.format)
                    styled_df['手当(円)'] = styled_df['手当(円)'].map('{:,.0f}'.format)
                    styled_df['合計(円)'] = styled_df['合計(円)'].map('{:,.0f}'.format)
                    
                    # データフレームの表示
                    st.markdown(f"### {name}様の精算データ")
                    st.dataframe(
                        styled_df,
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Excel用のデータフレームを準備（数値型を保持）
                    excel_df = person_data[['date', 'route', 'total_distance', 'transportation_fee', 'allowance', 'total']]
                    excel_df.columns = ['日付', '経路', '距離(km)', '交通費(円)', '手当(円)', '合計(円)']
                    
                    # Excel生成とダウンロードボタン
                    excel_data = create_expense_excel(excel_df, name)
                    st.download_button(
                        label=f"{name}様の清算書をダウンロード",
                        data=excel_data,
                        file_name=f"清算書_{name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == "__main__":
    main()
